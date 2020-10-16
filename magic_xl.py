'''
Excel formatting module v.1

October 2019
V.Nesterov

'''
# importing modules
import numpy as np
import pandas as pd
from datetime import datetime
from datetime import date

import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border, NamedStyle, Side, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
import string

def decor(path, file_name, tab_name, frame=[],
         built_in = False,
         path_target_file = None,
         row_h=None, col_ws=[20], header_h=30,
         conditional_coloring=None, exact_condition=True,
         conditional_first_row=False,
         merge = {},
         group_col = {},
         group_row = {},
         header_fill_color="DDDDDD", header_text_color='000000',
         body_color="FFFFFF",
         left_col_color="FFFFFF",
         highlight_row_dict={},
         highlight_index_dict={},
         hide_cols = [],
         freeze_top = False,
         footnote=None):
    '''
        *************************************************
        ***** Function to format Excel spreadsheets *****
        * .csv .xls or .xlsx format supported
        * Multiple tabs supported
        * Filters are set on top row automatically
        *************************************************
        Arguments:
            - path - full path to working directory (followed by '/');
            - file_name - target file name to save in (including extension);
                or desired file name if dataframe passed;
            - tab_name - desired tab name;
            - frame - pandas dataframe to work with (if not with file) - optional;
            - built_in - bool - True if to save with existing workbook
                                    (file_name will be used to return file).
            - path_target_file - string - if built in - full path with filename need to be provided.
            - row_h - row height, default=None - optional;
            - col_ws - columns width as a list of integers for each column in sequence,
                default=[20] - optional;
            - header_h - top row height, default=30 - optional;
            - conditional_coloring - dictionary where key is incell string to find and format and 
                value is color in HEX format (as string), default=None - optional;
            - exact_condition - boolean to exactly match keys in conditional formatting dictionary or contain the key in cell
            - conditional_first_row - boolean to catch 1st(header) row in conditional formatting setting
            - merge - dictionary - columns and rows dimensions to merge, e.g:
                {'key_col':[['row_start','row_end'], ...} where key_col and row_start/end are all integers.
            - group_row - dict where id key is unique identifier and value is list of grouping pair, e.g: {id:[1,3]}
            - group_col - dict where id key is unique identifier and value is list of grouping pair, e.g: {id:['A','B']} 
            - header_fill_color - color of top row in HEX (string), default="DDDDDD" (grey) - optional;
            - header_text_color - color of text in top row in HEX (string), default="000000" (black) - optional;
            - body_color - color of background in HEX (string), default="FFFFFF" (white) - optional;
            - left_col_color - color of left column in HEX (string), default="FFFFFF" (white) - optional;
            - highlight_row_dict- dictionary - to highlight rows - (row only highlighters) that consit of:
                id - integer - unique dictionary identifier for each highlighter
                rows - list of rows (integers)
                hlt_color - string - format: "bababa",
                hlt_txt_color - string - format: "000000";
                hlt_border_bold - boolean;
                hlt_font_bold - boolean;
                example: {1: {'rows':[1,3,4], 'hlt_color': "000000",
                                                  'hlt_txt_color':"000000",
                                                  'hlt_border_bold':True,
                                                  'hlt_font_bold':False}}
            - highlight_index_dict - dictionary - to highlight specific cells based on their index(position).
                Consist of:
                id - integer - unique dictionary identifier for each highlighter
                idxs - list of list pairs (integers) for unique cell address [row, column] 
                hlt_color - string - format: "bababa",
                hlt_txt_color - string - format: "000000";
                hlt_border_bold - boolean;
                hlt_font_bold - boolean;
                example: {1: {'idxs':[[1,3],[1,2],[3,4]],
                                                  'hlt_color': "000000",
                                                  'hlt_txt_color':"000000",
                                                  'hlt_border_bold':True,
                                                  'hlt_font_bold':False}}
                    
            - hide_cols - list - integers of columns(fields) to hide
            - freeze_top - bool - to freeze top row
            - footnote - footnote at the end of the document, default=None - optional.
        Returns: None
            Saves spreadhseet in same directory
    '''
    if path != "":
        if path[-1] != "/":
            path = path+"/"
    
    if len(merge) != 0:
        print("You are going to merge some columns.\n")
        
    # if converting file itself
    if len(frame) == 0:
        # opening file
        if ".xls" in file_name or ".xlsx" in file_name:
            data = pd.read_excel(path+file_name, header=0, dtype='object')
        elif ".csv" in file_name:
            data = pd.read_csv(path+file_name, dtype="object")
        else:
            print("Error: Wrong data format. Please supply only .csv .xls or .xlsx format.")
            return None
    else:
        # reading just dataframe
        data = frame.copy()
        
    # dimensions
    rows_no, cols_no = data.shape
    rows_no+=1
    
    # default columns dimensions             
    if cols_no != len(col_ws):
        print("Setting up default column width...\n"+\
              "If you want to specify width for each column, please \n"+\
              "specify it in col_ws list in order as columns follow.")
        col_ws = [20]*cols_no
        
    # creating letters/columns index
    l1 = list(string.ascii_uppercase)[:cols_no]
    l2, l3 = [],[]
    for i in l1:
        for l in l1:
            l2.append(i+l)
    for i in l1:
        for l in l2:
            l3.append(i+l)
    letters = l1+l2+l3
    letters = letters[:data.shape[1]]
    rows = rows_no + 1
        
    # checking datatypes
    dates_cols = {}
    for e, c in enumerate(data.columns):
        if np.issubdtype(data[c].dtype, np.datetime64):
            dates_cols[letters[e]] = 'mm-dd-yy'
        
    # dimensions
    rows_no, cols_no = data.shape
    rows_no+=1
    
    # ##### Filling in new workbook with data
    try:
        # Initializing openpyxl and setting up data
        if built_in:
            # using existing spreadsheet to augment data in
            wb = openpyxl.load_workbook(path_target_file)
            # create new sheet
            wb.create_sheet(tab_name)
            wb.active = -1
            # defining tab
            main_tab = wb.active
        else:
            # creating openpyxl object to read data in excel
            wb = Workbook()
            # defining tab
            main_tab = wb.active
            main_tab.title = tab_name
        # loading data
        for r in dataframe_to_rows(data, index=False, header=True):
            main_tab.append(r)
    except:
        # the reason of exception may be NaN/Nat values not convertable to objects
           # Initializing openpyxl and setting up data
        if built_in:
            # using existing spreadsheet to augment data in
            wb = openpyxl.load_workbook(path_target_file)
            # create new sheet
            wb.create_sheet(tab_name)
            wb.active = -1
            # defining tab
            main_tab = wb.active
        else:
            # creating openpyxl object to read data in excel
            wb = Workbook()
            # defining tab
            main_tab = wb.active
            main_tab.title = tab_name
        
        # replacing NaNs and loading data
        print("Failed convert dataframe to rows from first attempt. "+\
              "Replacing NaNs with '' to try again...")
        data.fillna("", inplace=True)
        for r in dataframe_to_rows(data, index=False, header=True):
            main_tab.append(r)
        print("Successfully converted dataframe to worksheet rows!")
    
    # #### Formatting
    print("Formatting \n")
    # ##### defining formatting styles
    # body style 
    def add_body_style(wb):
        name = 'body_' + str(tab_name)
        st = NamedStyle(name=name)
        st.font = Font(name='Calibri', bold=False, size=11)
        bd = Side(style='thin', color="000000")
        st.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        st.alignment=Alignment(horizontal='left',
                            vertical='center',
                            text_rotation=0,
                            wrap_text=True,
                            shrink_to_fit=False,
                            indent=0)
        st.fill = PatternFill(start_color=body_color,
                           end_color=body_color,
                           fill_type='solid')
        wb.add_named_style(st)
        return name
    
    # header style
    def add_head_style(wb):
        name = 'headstyle_' + str(tab_name)
        st = NamedStyle(name=name)
        st.font = Font(name='Calibri', bold=True, color=header_text_color, size=10)
        bd = Side(style='thin', color="000000")
        st.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        st.alignment=Alignment(horizontal='center',
                            vertical='center',
                            text_rotation=0,
                            wrap_text=True,
                            shrink_to_fit=False,
                            indent=0)
        st.fill = PatternFill(start_color=header_fill_color,
                           end_color=header_fill_color,
                           fill_type='solid')
        wb.add_named_style(st)
        return name
       
    # left column style
    def add_leftcol_style(wb):
        name = 'indexer_' + str(tab_name)
        st = NamedStyle(name=name)
        st.font = Font(name='Calibri', bold=False, size=10)
        bd = Side(style='thin', color="000000")
        st.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        st.alignment=Alignment(horizontal='left',
                            vertical='center',
                            text_rotation=0,
                            wrap_text=True,
                            shrink_to_fit=False,
                            indent=0)
        st.fill = PatternFill(start_color=left_col_color,
                           end_color=left_col_color,
                           fill_type='solid')
        wb.add_named_style(st)
        return name
        
    # highlighter style
    def add_row_hlt_style(wb, highlight_vals, name_order, h_type=None):
        '''
            This styler is for highlighting cells and works as
            for one style creation and for multiple as well due to 
            'name_order' argument which will distiguish each unique style.
            This function is used for both index and row highlighters
            Takes:
                - wb - object - workbook;
                - highlight_vals - dictionary from highlight_dict;
                - name_order - int - id from highlight_dict
                - h_type - string - highlighter type 'row' or 'index'
        '''
        if h_type == None:
            print("Error: add_row_hlt_style(): Please"+\
                  "specify correct h_type variable.")
            return None
        # create unique style names depending on type of highlight
        name = '{}_highlighter_'.format(h_type) + str(tab_name) + str(name_order)
        st = NamedStyle(name=name)
        st.font = Font(name='Calibri', bold=highlight_vals['hlt_font_bold'],
                       size=10, color=highlight_vals['hlt_txt_color'])
        if highlight_vals['hlt_border_bold']:
            bord_side = 'thick'
        else:
            bord_side = "thin"
        bdb = Side(style=bord_side, color="000000")
        bdt = Side(style='thin', color="000000")
        st.border = Border(left=bdt, top=bdb, right=bdt, bottom=bdb)
        st.alignment=Alignment(horizontal='left',
                            vertical='center',
                            text_rotation=0,
                            wrap_text=True,
                            shrink_to_fit=False,
                            indent=0)
        st.fill = PatternFill(start_color=highlight_vals['hlt_color'],
                           end_color=highlight_vals['hlt_color'],
                           fill_type='solid')
        wb.add_named_style(st)
        return name
    
    # merging cells 
    if len(merge) != 0:
        print("Merging rows\n")
        for k in merge:
            for row in merge[k]:
                main_tab.merge_cells(start_row=row[0],
                                     start_column=k,
                                     end_row=row[1],
                                     end_column=k)
    
    # ##### applying styles 
    # appending styles to workbook if new workbook
    # left column style
    index_style = add_leftcol_style(wb)
    for rw in range(2, rows):
        main_tab['A'+str(rw)].style = index_style
        # applying data types if any
        for k in dates_cols:
            if k == 'A':
                main_tab['A'+str(rw)].number_format = dates_cols[k]
                
    # header style
    head_style = add_head_style(wb)
    for l in letters:
        main_tab[l+"1"].style = head_style

    # body style
    body_style = add_body_style(wb)
    for rw in range(2, rows):
        for l in letters[1:]:
            main_tab[l+str(rw)].style = body_style
            # applying data types if any
            for k in dates_cols:
                if k == l:
                    main_tab[l+str(rw)].number_format = dates_cols[k]
                    
    # row highlighter style
    if len(highlight_row_dict) != 0:
        # sorting integer keys to apply in ascending order
        lst = list(highlight_row_dict.keys())
        lst.sort(reverse=False)
        for rw in lst:
            hlt_style = add_row_hlt_style(wb, highlight_row_dict[rw], rw,
                                                                      'row')
            for l in letters:
                for r in highlight_row_dict[rw]['rows']:
                    main_tab[l+str(r)].style = hlt_style
                    # applying data types if any
                    for k in dates_cols:
                        if k == l:
                            main_tab[l+str(r)].number_format = dates_cols[k]
    
    # index highlighter style
    if len(highlight_index_dict) != 0:
        # sorting integer keys to apply in ascending order
        lst = list(highlight_index_dict.keys())
        lst.sort(reverse=False)
        for i in lst:
            id_hlt_style = add_row_hlt_style(wb, highlight_index_dict[i], i,
                                                                      'index')
            r = highlight_index_dict[i]['idxs'][0]
            c = letters[highlight_index_dict[i]['idxs'][1]]
            main_tab[c+str(r)].style = id_hlt_style
            # applying data types if any
            for k in dates_cols:
                if k == c:
                    main_tab[c+str(r)].number_format = dates_cols[k]
                
    # ##### rows and columns dimensions
    # applying rows dimensions
    # header height
    main_tab.row_dimensions[1].height = header_h
    
    # regular row height
    if row_h != None:
        for dim in range(2, rows):
            main_tab.row_dimensions[dim].height = row_h
            
    # applying columns dimensions
    # iterating through number of columns
    for dim, w in zip(range(cols_no), col_ws):
        main_tab.column_dimensions[letters[dim]].width = w
    
    # ##### conditional formatting
    # adds conditional format to selected range - exact match
    def add_cond_text_format_exact(ws, text, color, start, end):
        '''
        Takes:
        - ws - worksheet object
        - text - as string
        - color - hex color
        - start cell+col string
        - end cell+col string
        '''
        fill = PatternFill(bgColor=color)
        dxf = DifferentialStyle(fill=fill)
        rule = Rule(type="cellIs", operator="equal", dxf=dxf)
        rule.formula = ['"{}"'.format(text)]
        ws.conditional_formatting.add(start+":"+end, rule)
    
    # adds conditional format to selected range - contains text
    def add_cond_text_format_contains(ws, text, color, start, end):
        '''
        Takes:
        - ws - worksheet object
        - text - as string
        - color - hex color
        - start cell+col string
        - end cell+col string
        '''
        print("using non-exact cond formatting")
        fill = PatternFill(bgColor=color)
        dxf = DifferentialStyle(fill=fill)
        rule = Rule(type="containsText", operator="containsText", text=text, dxf=dxf)
        rule.formula = ['NOT(ISERROR(SEARCH("{}",A2)))'.format(text)]
        ws.conditional_formatting.add(start+":"+end, rule)
    
    # inserting conditional formatting formula for ratings
    if conditional_coloring != None:
        if conditional_first_row:
            cell_start = "A1"
        else:
            cell_start = "A2"
        if exact_condition:
            for val in conditional_coloring:
                add_cond_text_format_exact(main_tab, val,
                                           conditional_coloring[val],
                                           cell_start, letters[-1]+str(rows_no))
        else:
            for val in conditional_coloring:
                add_cond_text_format_contains(main_tab, val,
                                           conditional_coloring[val],
                                           cell_start, letters[-1]+str(rows_no))

    # ##### making filters
    print("Finalizing \n")
    # filtering
    main_tab.auto_filter.ref = "A1:"+letters[-1]+"1"
    
    # ##### other useful stuff
    # hiding gridlines
    main_tab.sheet_view.showGridLines = False
    
    # freezing panes
    if freeze_top:
        main_tab.freeze_panes = "A2"
        
    # grouping columns
    if len(group_col) != 0:
       for cg in group_col:
           main_tab.column_dimensions.group(group_col[cg][0],
                                            group_col[cg][1],
                                            hidden=True)
    
    # grouping rows
    if len(group_row) != 0:
       for rg in group_row:
           main_tab.row_dimensions.group(group_row[rg][0],
                                            group_row[rg][1],
                                            hidden=True)
    
    # hide columns
    if len(hide_cols) != 0:
        for hc in hide_cols:
           main_tab.column_dimensions[letters[hc]].hidden = True 
    
    # putting reference at the end of document
    if footnote != None:
        main_tab['A'+str(rows_no+2)].value = '* '+footnote
        
    # Saving workbook
    if "_fmt.xlsx" not in file_name:
        if "." in file_name:
            file_name = file_name.split(".")[0]+"_fmt.xlsx"
            file_name = file_name.replace(" ","_")
        elif "." not in file_name:
            file_name = file_name+"_fmt.xlsx"
            file_name = file_name.replace(" ","_")
    wb.save(filename = path+file_name)
    wb.close()  
    print("Finished formatting, saved.")
    
